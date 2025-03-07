import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import os
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import OneHotEncoder
import numpy as np

st.title("DataVerify")

# Function to build validation mappings from the template
def build_validation_mappings(sheet):
    """
    Parse the Excel sheet to extract dropdown values and build VALIDATION_MAPPINGS.
    
    Args:
        sheet: openpyxl Worksheet object
    
    Returns:
        dict: VALIDATION_MAPPINGS with material-specific forms and details
    """
    # Find the row with dropdown headers
    for row in sheet.iter_rows(min_row=1, max_row=20):
        headers = [cell.value for cell in row if cell.value]
        if any("Packaging Form List" in str(header) for header in headers):
            dropdown_header_row = row.row
            break
    else:
        st.error("Could not find dropdown headers in the template.")
        st.stop()

    # Get headers from the dropdown row
    headers = [cell.value for cell in sheet[dropdown_header_row]]

    # Map materials to their form and detail columns
    material_forms = {}
    material_details = {}
    for col_idx, header in enumerate(headers, start=1):
        if header and "Packaging Form List" in header:
            material = header.replace(" Packaging Form List", "")
            material_forms[material] = col_idx
        elif header and "Further Details" in header:
            material = header.replace(" Further Details", "")
            material_details[material] = col_idx

    # Build VALIDATION_MAPPINGS
    VALIDATION_MAPPINGS = {}
    for material in material_forms:
        # Read Packaging Form list
        forms_col = material_forms[material]
        forms_list = []
        for row in sheet.iter_rows(min_row=dropdown_header_row + 1, 
                                 max_col=forms_col, 
                                 min_col=forms_col, 
                                 values_only=True):
            value = row[0]
            if value is None:
                break
            forms_list.append(value)

        # Read Further Details list (if exists)
        details_col = material_details.get(material)
        details_list = []
        if details_col:
            for row in sheet.iter_rows(min_row=dropdown_header_row + 1, 
                                     max_col=details_col, 
                                     min_col=details_col, 
                                     values_only=True):
                value = row[0]
                if value is None:
                    break
                details_list.append(value)

        VALIDATION_MAPPINGS[material] = {
            "Packaging Form": forms_list,
            "Further Details": details_list
        }

    return VALIDATION_MAPPINGS

# Function to save to historical data
def save_to_historical_data(df, file_path="historical_data.csv"):
    """Append validated data to historical CSV, ensuring no NaN in Weight (kg)."""
    df_subset = df[["Packaging Material", "Packaging Form", "Further Details", "Weight (kg)"]]
    df_subset = df_subset.dropna(subset=["Weight (kg)"])
    if os.path.exists(file_path):
        df_subset.to_csv(file_path, mode='a', header=False, index=False)
    else:
        df_subset.to_csv(file_path, mode='w', header=True, index=False)

# Function to train the anomaly detection model
def train_anomaly_model(file_path="historical_data.csv"):
    """Train RandomForestRegressor, handling NaN values in historical data."""
    if not os.path.exists(file_path):
        return None, None, None
    
    historical_df = pd.read_csv(file_path)
    if historical_df.empty:
        return None, None, None
    
    features = ["Packaging Material", "Packaging Form", "Further Details"]
    target = "Weight (kg)"
    historical_df = historical_df.dropna(subset=[target] + features)
    if historical_df.empty:
        return None, None, None
    
    X = historical_df[features]
    y = historical_df[target]
    encoder = OneHotEncoder(sparse_output=False, handle_unknown='ignore')
    X_encoded = encoder.fit_transform(X)
    model = RandomForestRegressor(n_estimators=100, random_state=42)
    model.fit(X_encoded, y)
    y_pred = model.predict(X_encoded)
    errors = np.abs(y - y_pred)
    error_threshold = np.percentile(errors, 95)
    
    return model, encoder, error_threshold

# Function to validate data against the ML model
def validate_against_ml_model(df, model, encoder, error_threshold):
    """Check for anomalies using the trained model."""
    if model is None or encoder is None or error_threshold is None:
        return ["No historical data available for ML validation."]
    
    warnings = []
    features = ["Packaging Material", "Packaging Form", "Further Details"]
    target = "Weight (kg)"
    
    for idx, row in df.iterrows():
        material = row["Packaging Material"]
        form = row["Packaging Form"]
        details = row["Further Details"]
        weight = row["Weight (kg)"]
        
        if pd.notna(material) and pd.notna(form) and pd.notna(weight):
            input_df = pd.DataFrame([[material, form, details]], columns=features)
            X_encoded = encoder.transform(input_df)
            predicted_weight = model.predict(X_encoded)[0]
            error = abs(weight - predicted_weight)
            
            if error > error_threshold:
                warnings.append(
                    f"Row {idx + 1}: Weight {weight} kg for '{material}' + '{form}' + '{details}' "
                    f"deviates significantly from predicted {predicted_weight:.2f} kg (error: {error:.2f} kg)."
                )
    
    return warnings

# Function to validate data structure and content
def validate_data(df, validation_mappings):
    """Validate data against dynamic validation mappings."""
    errors = []
    expected_columns = ["packaging material", "packaging form", "further details", "weight (kg)"]
    actual_columns = [col.lower().strip() if isinstance(col, str) else col for col in df.columns]
    column_mapping = {}
    
    for expected in expected_columns:
        for actual in actual_columns:
            if expected in actual:
                column_mapping[expected] = df.columns[actual_columns.index(actual)]
                break
        if expected not in column_mapping:
            errors.append(f"Missing required column: {expected.capitalize()}")
            return errors

    material_col = column_mapping["packaging material"]
    form_col = column_mapping["packaging form"]
    details_col = column_mapping["further details"]
    weight_col = column_mapping["weight (kg)"]

    for idx, row in df.iterrows():
        material = row[material_col]
        form = row[form_col]
        details = row[details_col]
        weight = row[weight_col]

        if all(pd.isna(row[col]) for col in [material_col, form_col, details_col, weight_col]):
            continue

        if pd.notna(weight):
            if isinstance(weight, (int, float)) and weight < 0:
                errors.append(f"Row {idx + 1}: Weight must be a positive number.")
            elif not isinstance(weight, (int, float)):
                errors.append(f"Row {idx + 1}: Weight must be a number, got '{weight}'.")
        elif pd.notna(material) and pd.notna(form):
            errors.append(f"Row {idx + 1}: Weight (kg) is required when Packaging Material and Packaging Form are provided.")

        if pd.isna(material) and any(pd.notna(row[col]) for col in [form_col, details_col, weight_col]):
            errors.append(f"Row {idx + 1}: Packaging Material is required when other fields are provided.")
            continue

        if pd.isna(material):
            continue

        if material not in validation_mappings:
            errors.append(f"Row {idx + 1}: Invalid Packaging Material '{material}'.")
            continue

        form = form.strip() if isinstance(form, str) else form
        details = details.strip() if isinstance(details, str) else details

        valid_forms = validation_mappings[material]["Packaging Form"]
        if pd.notna(form) and form not in valid_forms:
            errors.append(f"Row {idx + 1}: '{form}' is not a valid Packaging Form for '{material}'. Valid options: {', '.join(valid_forms)}")

        valid_details = validation_mappings[material]["Further Details"]
        if pd.notna(details):
            if material == "Wood" and details == "N/A":
                continue
            if not valid_details:
                errors.append(f"Row {idx + 1}: No Further Details should be specified for '{material}' (except 'N/A' for Wood).")
            elif details not in valid_details:
                errors.append(f"Row {idx + 1}: '{details}' is not a valid Further Detail for '{material}'. Valid options: {', '.join(valid_details)}")

    return errors

# File uploader and main logic
uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
if uploaded_file is not None:
    # Load workbook with data_only=True to get calculated values
    wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
    packaging_sheet = wb["Packaging Data"]

    # Build dynamic validation mappings
    VALIDATION_MAPPINGS = build_validation_mappings(packaging_sheet)

    # Find data header row
    for row_num, row in enumerate(packaging_sheet.iter_rows(min_row=1, values_only=True), start=1):
        if len(row) >= 4 and row[0] == "Packaging Material" and row[1] == "Packaging Form" and \
           row[2] == "Further Details" and row[3] == "Weight (kg)":
            data_header_row = row_num
            break
    else:
        st.error("Could not find data headers in the template.")
        st.stop()

    # Load data into DataFrame
    data = []
    for row in packaging_sheet.iter_rows(min_row=data_header_row + 1, values_only=True):
        data.append(row)
    headers = [cell.value for cell in packaging_sheet[data_header_row]]
    df = pd.DataFrame(data, columns=headers)

    key_columns = ["Packaging Material", "Packaging Form", "Further Details", "Weight (kg)"]
    filtered_df = df.dropna(subset=key_columns, how='all')

    st.write("Original Data:")
    st.dataframe(filtered_df)

    st.write("Edit the data below:")
    edited_df = st.data_editor(filtered_df, num_rows="dynamic", use_container_width=True)

    # Train the model
    model, encoder, error_threshold = train_anomaly_model()

    # Validate
    errors = validate_data(edited_df, VALIDATION_MAPPINGS)
    historical_warnings = validate_against_ml_model(edited_df, model, encoder, error_threshold)

    if errors:
        st.error("Validation Errors Found:")
        for error in errors:
            st.write(f"- {error}")
    else:
        st.success("Data is valid!")

    if historical_warnings:
        st.warning("Suspicious Outlier Warnings:")
        for warning in historical_warnings:
            st.write(f"- {warning}")

    if st.button("Save to Excel"):
        save_to_historical_data(edited_df)

        # Clear existing data in the sheet
        for row in packaging_sheet.iter_rows(min_row=data_header_row + 1, 
                                           max_row=packaging_sheet.max_row, 
                                           max_col=len(headers)):
            for cell in row:
                cell.value = None

        # Write edited data back
        for index, row in edited_df.iterrows():
            for col_idx, value in enumerate(row):
                packaging_sheet.cell(row=data_header_row + index + 1, column=col_idx + 1, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="Download Updated Excel",
            data=output,
            file_name="updated_file.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.write("Instructions: Upload an Excel file, edit the data, and click 'Save to Excel' to download the updated file.")
