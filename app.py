import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import os

st.title("GreenAccuracy Editor")

# Define the mappings (unchanged)
VALIDATION_MAPPINGS = {
    "Plastic": {
        "Packaging Form": [
            "Beverage bottle", "Carrier bag", "Disposables (bowls, boxes, covers, cups, plates, trays)",
            "Product packaging (flexible)", "Product packaging (rigid, excluding beverage bottle)",
            "Transport and protective packaging", "Others", "Crate"
        ],
        "Further Details": ["EPS", "HDPE", "LDPE", "PET", "PP", "PS", "PVC", "Others"]
    },
    "Paper": {
        "Packaging Form": [
            "Carrier bag", "Disposables (bowls, boxes, cups, plates, trays)",
            "Product packaging", "Transport and protective packaging", "Others"
        ],
        "Further Details": ["Corrugated board", "Paper", "Paperboard", "Others"]
    },
    "Metal": {
        "Packaging Form": [
            "Beverage can", "Disposables (foils, trays)", "Food can",
            "Product packaging (excluding beverage can, food can)", "Others"
        ],
        "Further Details": ["Aluminium", "Steel", "Tin", "Others"]
    },
    "Glass": {
        "Packaging Form": ["Beverage bottle", "Product packaging (excluding beverage bottle)", "Others"],
        "Further Details": ["Brown", "Clear", "Green", "Other colours"]
    },
    "Wood": {
        "Packaging Form": ["Crate", "Pallet", "Product packaging", "Transport and protective packaging", "Others"],
        "Further Details": ["N/A"]
    },
    "Composite": {
        "Packaging Form": ["Beverage carton", "Packs / Sachets", "Product packaging (excluding beverage carton, packs/sachets)", "Others"],
        "Further Details": []
    },
    "Others": {
        "Packaging Form": ["Carrier bag", "Others"],
        "Further Details": ["Biodegradable/compostable", "Oxo-degradable/oxo-biodegradable", "Others"]
    }
}

# Function to save to historical data
def save_to_historical_data(df, file_path="historical_data.csv"):
    df_subset = df[["Packaging Material", "Packaging Form", "Further Details", "Weight (kg)"]]
    if os.path.exists(file_path):
        df_subset.to_csv(file_path, mode='a', header=False, index=False)
    else:
        df_subset.to_csv(file_path, mode='w', header=True, index=False)

# Function to get historical ranges
def get_historical_ranges(file_path="historical_data.csv"):
    if not os.path.exists(file_path):
        return {}
    historical_df = pd.read_csv(file_path)
    ranges = historical_df.groupby(["Packaging Material", "Packaging Form"])["Weight (kg)"].agg(['min', 'max']).to_dict('index')
    range_dict = {}
    for (material, form), stats in ranges.items():
        range_dict[(material, form)] = (stats['min'], stats['max'])
    return range_dict

# Function to validate against historical ranges
def validate_against_historical(df, historical_ranges):
    warnings = []
    for idx, row in df.iterrows():
        material = row["Packaging Material"]
        form = row["Packaging Form"]
        weight = row["Weight (kg)"]
        
        if pd.notna(material) and pd.notna(form) and pd.notna(weight):
            key = (material, form)
            if key in historical_ranges:
                min_weight, max_weight = historical_ranges[key]
                if isinstance(weight, (int, float)) and (weight < min_weight or weight > max_weight):
                    warnings.append(
                        f"Row {idx + 1}: Weight {weight} kg for '{material}' + '{form}' "
                        f"is outside historical range ({min_weight}–{max_weight} kg)."
                    )
    return warnings

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
if uploaded_file is not None:
    # Load the workbook and sheets
    wb = load_workbook(io.BytesIO(uploaded_file.getvalue()))
    packaging_sheet = wb["Packaging Data"]

    header_row = 6
    headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in packaging_sheet[header_row + 1]]
    data = []
    for row in packaging_sheet.iter_rows(min_row=header_row + 2, values_only=True):
        data.append(row)
    df = pd.DataFrame(data, columns=headers)

    # Filter out empty rows (all key columns NaN)
    key_columns = ["Packaging Material", "Packaging Form", "Further Details", "Weight (kg)"]
    filtered_df = df.dropna(subset=key_columns, how='all')

    st.write("Original Data:")
    st.dataframe(filtered_df)

    st.write("Edit the data below:")
    edited_df = st.data_editor(filtered_df, num_rows="dynamic", use_container_width=True)

    # Validation Logic (existing)
    def validate_data(df):
        errors = []
        expected_columns = ["packaging material", "packaging form", "further details", "weight (kg)"]
        actual_columns = [col.lower().strip() if isinstance(col, str) else col for col in df.columns]
        column_mapping = {}
        for expected in expected_columns:
            for actual in actual_columns:
                if expected in actual:
                    column_mapping[expected] = df.columns[actual_columns.index(actual)]
                    break
            if expected not in column_mapping:
                errors.append(f"Missing required column: {expected.capitalize()}")
                return errors

        material_col = column_mapping["packaging material"]
        form_col = column_mapping["packaging form"]
        details_col = column_mapping["further details"]
        weight_col = column_mapping["weight (kg)"]

        for idx, row in df.iterrows():
            material = row[material_col]
            form = row[form_col]
            details = row[details_col]
            weight = row[weight_col]

            if all(pd.isna(row[col]) for col in [material_col, form_col, details_col, weight_col]):
                continue

            if pd.notna(weight):
                if isinstance(weight, (int, float)) and weight < 0:
                    errors.append(f"Row {idx + 1}: Weight must be a positive number.")
                elif not isinstance(weight, (int, float)):
                    errors.append(f"Row {idx + 1}: Weight must be a number, got '{weight}'.")

            if pd.isna(material) and any(pd.notna(row[col]) for col in [form_col, details_col, weight_col]):
                errors.append(f"Row {idx + 1}: Packaging Material is required when other fields are provided.")
                continue

            if pd.isna(material):
                continue

            if material not in VALIDATION_MAPPINGS:
                errors.append(f"Row {idx + 1}: Invalid Packaging Material '{material}'.")
                continue

            form = form.strip() if isinstance(form, str) else form
            details = details.strip() if isinstance(details, str) else details

            valid_forms = VALIDATION_MAPPINGS[material]["Packaging Form"]
            if pd.notna(form) and form not in valid_forms:
                errors.append(f"Row {idx + 1}: '{form}' is not a valid Packaging Form for '{material}'. Valid options: {', '.join(valid_forms)}")

            valid_details = VALIDATION_MAPPINGS[material]["Further Details"]
            if pd.notna(details):
                if material == "Wood" and details == "N/A":
                    continue
                if not valid_details:
                    errors.append(f"Row {idx + 1}: No Further Details should be specified for '{material}' (except 'N/A' for Wood).")
                elif details not in valid_details:
                    errors.append(f"Row {idx + 1}: '{details}' is not a valid Further Detail for '{material}'. Valid options: {', '.join(valid_details)}")

        return errors

    # Perform validations
    errors = validate_data(edited_df)
    historical_ranges = get_historical_ranges()
    historical_warnings = validate_against_historical(edited_df, historical_ranges)

    if errors:
        st.error("Validation Errors Found:")
        for error in errors:
            st.write(f"- {error}")
    else:
        st.success("Data is valid!")

    if historical_warnings:
        st.warning("Historical Range Warnings:")
        for warning in historical_warnings:
            st.write(f"- {warning}")

    # Save button and historical data update
    if st.button("Save to Excel"):
        # Save to historical data
        save_to_historical_data(edited_df)

        # Clear existing data in sheet
        for row in packaging_sheet.iter_rows(min_row=header_row + 2, max_row=packaging_sheet.max_row, max_col=len(headers)):
            for cell in row:
                cell.value = None

        # Write new data (using edited_df, which is based on filtered_df)
        for index, row in edited_df.iterrows():
            for col_idx, value in enumerate(row):
                packaging_sheet.cell(row=header_row + index + 2, column=col_idx + 1, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        st.download_button(
            label="Download Updated Excel",
            data=output,
            file_name="updated_file.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.write("Instructions: Upload an Excel file, edit the data, and click 'Save to Excel' to download the updated file.")
